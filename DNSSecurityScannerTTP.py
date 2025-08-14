#!/usr/bin/env python3
"""
DNS Security Scanner (SPF / DMARC / DKIM) — MSP-ready

What it does
------------
- Pulls SPF, DMARC, and (common) DKIM selector DNS records for each domain.
- Flags misconfigurations similar to popular scanners (e.g., weak DMARC policy, SPF over 10 lookups, DKIM key too short).
- Produces a **results.xlsx** that mirrors EasyDMARC-style findings plus your CRM columns.

How to run
----------
# Put an Excel file named 'domains.xlsx' next to this script with a column header 'Domain'
python DNSSecurityScannerTTP.py

Outputs:
- results.xlsx (nicely formatted, includes your CRM header fields)
- results.csv (flat export of the core findings)
"""

import json
import requests
import argparse
import base64
import concurrent.futures as futures
import csv
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

# ---------- Dependencies ----------
try:
    import pandas as pd
except Exception:
    print("Missing dependency: pandas. Install with: pip install pandas", file=sys.stderr)
    raise

try:
    import dns.resolver
    import dns.exception
    import dns.name
except ImportError:
    print("Missing dependency: dnspython. Install with: pip install dnspython", file=sys.stderr)
    raise

# Try to use cryptography for accurate DKIM RSA length detection (optional)
try:
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.backends import default_backend
    HAS_CRYPTO = True
except Exception:
    HAS_CRYPTO = False

# --------------------------- Configuration ---------------------------

PUBLIC_RESOLVERS = ["1.1.1.1", "8.8.8.8"]
TIMEOUT = 3.0
LIFETIME = 4.0
MAX_WORKERS = 20

# Common DKIM selectors used by major ESPs and platforms.
SELECTOR_WORDLIST = [
    # Generic/common
    "default", "selector", "selector1", "selector2", "s1", "s2", "k1", "k2",
    # Microsoft 365
    "selector1", "selector2",
    # Google Workspace
    "google",
    # SendGrid
    "s1", "s2",
    # Mailchimp / Mandrill
    "k1", "k2", "mandrill", "mandrill1", "mandrill2",
    # Zoho
    "zoho", "zohomail",
    # Fastmail
    "fm1", "fm2", "fm3",
    # AWS SES
    "amazon", "amazonses", "ses", "mail",
    # SparkPost
    "scph", "scph1", "scph2",
    # Postmark
    "pm", "pm-bounces",
    # Mailgun
    "smtp", "mg", "mailo", "mta",
    # Brevo (Sendinblue)
    "sib", "mail", "mta1", "mta2",
    # HubSpot
    "hs1", "hs2",
    # Klaviyo
    "krs", "kl",
    # Salesforce Marketing Cloud
    "sfdc", "sfmc1", "sfmc2",
    # Campaign Monitor
    "cm", "cm1", "cm2",
    # Generic extras
    "mail", "email", "smtp", "news", "n1", "n2"
]

# --------------------------- Utilities ---------------------------

def _resolver_chain() -> List[dns.resolver.Resolver]:
    """Yield resolvers in order: system → public(Cloudflare/Google)."""
    res = []

    # System resolver (respects /etc/resolv.conf, corporate DNS, split-horizon, etc.)
    try:
        r_sys = dns.resolver.Resolver(configure=True)
        r_sys.timeout = TIMEOUT
        r_sys.lifetime = LIFETIME
        res.append(r_sys)
    except Exception:
        pass

    # Public resolvers (Cloudflare, Google)
    try:
        r_pub = dns.resolver.Resolver(configure=False)
        r_pub.nameservers = PUBLIC_RESOLVERS
        r_pub.timeout = TIMEOUT
        r_pub.lifetime = LIFETIME
        res.append(r_pub)
    except Exception:
        pass

    return res

def _normalize_txt_piece(s: str) -> str:
    # Strip leading/trailing quotes and collapse whitespace between quoted chunks
    s = s.strip()
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1]
    return s

def _merge_txt_strings(rr) -> str:
    """Robustly extract TXT RDATA for dnspython 2.x, merging split strings."""
    # Prefer rr.strings if available (older dnspython). For newer, use to_text parsing.
    try:
        if hasattr(rr, 'strings') and rr.strings:
            return "".join([b.decode("utf-8", errors="replace") for b in rr.strings])
    except Exception:
        pass
    text = rr.to_text()  # e.g. "\"v=DMARC1; p=reject;\" \"rua=mailto:...\""
    # Extract quoted segments and join
    parts, cur, quoted = [], "", False
    for ch in text:
        if ch == '"':
            quoted = not quoted
            if not quoted:
                parts.append(cur)
                cur = ""
        elif quoted:
            cur += ch
    if parts:
        return "".join(parts)
    return _normalize_txt_piece(text)

def _doh_txt_records(name: str) -> List[str]:
    """DNS-over-HTTPS fallback via Google DoH."""
    try:
        url = f"https://dns.google/resolve?name={name}&type=TXT"
        resp = requests.get(url, timeout=4)
        resp.raise_for_status()
        data = resp.json()
        out: List[str] = []
        for ans in data.get("Answer", []) or []:
            if ans.get("type") == 16:  # TXT
                raw = ans.get("data", "")
                parts, cur, quoted = [], "", False
                for ch in raw:
                    if ch == '"':
                        quoted = not quoted
                        if not quoted:
                            parts.append(cur)
                            cur = ""
                    elif quoted:
                        cur += ch
                if parts:
                    out.append("".join(parts).strip())
                else:
                    out.append(_normalize_txt_piece(raw))
        return out
    except Exception:
        return []

def txt_records(name: str) -> List[str]:
    """Return TXT strings for a DNS name (without quotes), using multiple resolvers + DoH fallback."""
    out: List[str] = []

    for r in _resolver_chain():
        try:
            answers = r.resolve(name, "TXT", lifetime=LIFETIME)
            for rr in answers:
                s = _merge_txt_strings(rr)
                if s:
                    out.append(s)
            if out:
                return out
        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.resolver.NoNameservers, dns.exception.Timeout):
            continue
        except Exception:
            continue

    doh_out = _doh_txt_records(name)
    if doh_out:
        return doh_out

    return out

def first_txt(name: str) -> Optional[str]:
    recs = txt_records(name)
    return recs[0] if recs else None

# --------------------------- SPF ---------------------------

SPF_MECH_REGEX = re.compile(r"(?P<mech>\b(?:a|mx|ptr|include|exists)\b)(?::(?P<arg>[^ \t]+))?|redirect=(?P<redir>[^ \t;]+)")
SPF_ALL_REGEX = re.compile(r"(?P<qual>[\+\-\~\?]?)all\b")

def get_spf(domain: str) -> Tuple[Optional[str], List[str], Dict[str, str]]:
    """Return SPF record, warnings, parsed tags."""
    warnings = []
    spf_records = [t for t in txt_records(domain) if t.lower().startswith("v=spf1")]
    if not spf_records:
        return None, ["No SPF record"], {}
    if len(spf_records) > 1:
        warnings.append("Multiple SPF records found (should be single record)")
    spf = spf_records[0]

    parsed: Dict[str, str] = {}
    lookup_count = count_spf_lookups(spf, visited=set())

    if lookup_count > 10:
        warnings.append(f"SPF exceeds 10 DNS-mechanism lookups (={lookup_count})")

    all_match = SPF_ALL_REGEX.search(spf)
    if all_match:
        qual = all_match.group("qual") or "+"
        if qual in ("+", "?"):
            warnings.append(f"Weak SPF qualifier on 'all' ({qual}all). Prefer ~all or -all.")
    else:
        warnings.append("No 'all' mechanism found; consider ~all or -all at end")

    if " ptr" in f" {spf.lower()} ":
        warnings.append("Uses 'ptr' mechanism (deprecated)")

    parsed["lookup_count"] = str(lookup_count)
    return spf, warnings, parsed

def count_spf_lookups(spf: str, visited: set, depth: int = 0) -> int:
    """Roughly count SPF DNS-mechanism lookups including include/redirect recursion."""
    if depth > 20:
        return 0
    count = 0
    for m in SPF_MECH_REGEX.finditer(spf):
        mech = m.group("mech")
        arg = m.group("arg")
        redir = m.group("redir")
        if redir:
            tgt = redir.strip()
            if tgt not in visited:
                visited.add(tgt)
                spf2 = first_txt(tgt)
                if spf2 and spf2.lower().startswith("v=spf1"):
                    count += 1 + count_spf_lookups(spf2, visited, depth + 1)
                else:
                    count += 1
        elif mech in ("include", "a", "mx", "ptr", "exists"):
            count += 1
            if mech == "include" and arg:
                tgt = arg.strip()
                if tgt not in visited:
                    visited.add(tgt)
                    spf2s = [t for t in txt_records(tgt) if t.lower().startswith("v=spf1")]
                    if spf2s:
                        count += count_spf_lookups(spf2s[0], visited, depth + 1)
    return count

# --------------------------- DMARC ---------------------------

def parse_kv(record: str) -> Dict[str, str]:
    """Parse DMARC-like tag=value;tag2=value2 string into dict (lowercased keys)."""
    out = {}
    for piece in record.split(";"):
        piece = piece.strip()
        if not piece:
            continue
        if "=" in piece:
            k, v = piece.split("=", 1)
            out[k.strip().lower()] = v.strip()
    return out

def get_dmarc(domain: str) -> Tuple[Optional[str], List[str], Dict[str, str]]:
    name = f"_dmarc.{domain}"
    rec = first_txt(name)
    warnings: List[str] = []
    parsed: Dict[str, str] = {}
    if not rec or "v=DMARC1" not in rec.upper():
        return None, ["No DMARC record"], {}

    rec_norm = re.sub(r"\s+", " ", rec).strip()
    tags = parse_kv(rec_norm)

    if tags.get("v", "").upper() != "DMARC1":
        warnings.append("DMARC version not 'DMARC1'")
    policy = tags.get("p", "").lower()
    if policy in ("", "none"):
        warnings.append("Weak DMARC policy (p=none). Consider p=quarantine or p=reject.")
    elif policy not in ("reject", "quarantine", "none"):
        warnings.append(f"Unknown DMARC policy (p={policy})")

    aspf = tags.get("aspf", "r").lower()
    adkim = tags.get("adkim", "r").lower()
    if aspf != "s" or adkim != "s":
        warnings.append("Relaxed alignment (aspf/adkim). Consider strict (s) if possible.")

    rua = tags.get("rua", "")
    ruf = tags.get("ruf", "")
    if not rua:
        warnings.append("No aggregate reporting URI (rua) configured")

    pct = tags.get("pct", "100")
    try:
        pct_val = int(pct)
        if pct_val < 100 and policy in ("quarantine", "reject"):
            warnings.append(f"DMARC applied to only pct={pct_val}% of mail")
    except ValueError:
        warnings.append(f"Invalid pct value ({pct})")

    sp = tags.get("sp", "")
    if sp and sp in ("none",):
        warnings.append("Subdomain policy sp=none (weak)")

    parsed.update({
        "policy": policy,
        "aspf": aspf,
        "adkim": adkim,
        "rua": rua,
        "ruf": ruf,
        "pct": tags.get("pct", "100"),
        "sp": sp or "",
        "fo": tags.get("fo", ""),
    })
    return rec_norm, warnings, parsed

# --------------------------- DKIM ---------------------------

def parse_dkim_txt(txt: str) -> Dict[str, str]:
    return parse_kv(txt)

def estimate_key_bits_from_p(p_b64: str) -> Optional[int]:
    """Estimate DKIM RSA key bits from base64-encoded 'p'."""
    try:
        raw = base64.b64decode(p_b64 + "==", validate=False)
        if not raw:
            return None
        if HAS_CRYPTO:
            try:
                pubkey = serialization.load_der_public_key(raw, backend=default_backend())
                if hasattr(pubkey, "key_size"):
                    return int(pubkey.key_size)
            except Exception:
                pass
        return len(raw) * 8
    except Exception:
        return None

def try_dkim_selector(domain: str, selector: str) -> Optional[Dict[str, str]]:
    name = f"{selector}._domainkey.{domain}"
    txts = txt_records(name)
    for t in txts:
        if "v=DKIM1" in t.upper():
            return parse_dkim_txt(t)
    return None

def get_dkim(domain: str, selectors: List[str]) -> Tuple[List[str], List[str], List[Dict[str, str]]]:
    """Return (warnings, found_selectors, records)."""
    warnings: List[str] = []
    found: List[str] = []
    recs: List[Dict[str, str]] = []

    for sel in selectors:
        rec = try_dkim_selector(domain, sel)
        if rec:
            found.append(sel)
            recs.append(rec)

    if not found:
        warnings.append("No DKIM records found using common selectors")
    else:
        for sel, rec in zip(found, recs):
            k = rec.get("k", "rsa").lower()
            p = rec.get("p", "")
            t = rec.get("t", "")
            s = rec.get("s", "")
            notes = []
            if k != "rsa":
                notes.append(f"Non-RSA key type ({k})")
            bits = estimate_key_bits_from_p(p) if p else None
            if bits is not None and bits < 1024:
                notes.append(f"Weak DKIM key length ({bits} bits)")
            if "s" in t:
                notes.append("t=s (strict) — only signed subdomains allowed")
            if s:
                notes.append(f"Service types in 's' tag: {s}")
            if notes:
                warnings.append(f"Selector '{sel}': " + "; ".join(notes))

    return warnings, found, recs

# --------------------------- Scoring & Status ---------------------------

def classify_status(present: bool, warnings_text: str) -> str:
    if not present:
        return "Invalid"
    if warnings_text.strip():
        return "Warning"
    return "Valid"

def score_row(spf_present: bool, spf_rec: str, spf_warn: str,
              dmarc_present: bool, dmarc_policy: str, dmarc_warn: str,
              dkim_found: bool, dkim_warn: str) -> int:
    score = 0
    # SPF: +2 if present, +1 if has -all or ~all and no major SPF warnings beyond alignment note
    if spf_present:
        score += 2
        if (" -all" in f" {spf_rec} ") or (" ~all" in f" {spf_rec} "):
            score += 1
    # DMARC: +1 if present at all, +3 if policy is quarantine/reject, +0 if p=none
    if dmarc_present:
        score += 1
        if dmarc_policy in ("quarantine", "reject"):
            score += 3
    # DKIM: +2 if any selector found, +1 if no warnings about weak keys
    if dkim_found:
        score += 2
        if not dkim_warn.strip():
            score += 1
    return max(0, min(score, 10))

def risk_from_score(score: int) -> str:
    if score <= 3:
        return "CRITICAL"
    if 4 <= score <= 6:
        return "MEDIUM"
    if 7 <= score <= 8:
        return "HIGH"
    return "LOW"  # 9–10

# --------------------------- Orchestration ---------------------------

def scan_domain(domain: str) -> Dict[str, str]:
    domain = domain.strip().lower().rstrip(".")
    spf_rec, spf_w, spf_parsed = get_spf(domain)
    dmarc_rec, dmarc_w, dmarc_parsed = get_dmarc(domain)
    dkim_w, dkim_found, _dkim_recs = get_dkim(domain, SELECTOR_WORDLIST)

    spf_present = bool(spf_rec)
    dmarc_present = bool(dmarc_rec)

    spf_status = classify_status(spf_present, " | ".join(spf_w))
    dmarc_status = classify_status(dmarc_present, " | ".join(dmarc_w))
    dkim_status = "Valid" if dkim_found and not dkim_w else ("Warning" if dkim_found else "Invalid")

    score = score_row(
        spf_present, spf_rec or "", " | ".join(spf_w),
        dmarc_present, (dmarc_parsed.get("policy", "") if dmarc_parsed else ""), " | ".join(dmarc_w),
        bool(dkim_found), " | ".join(dkim_w),
    )
    risk_lvl = risk_from_score(score)

    # Narrative summary similar to EasyDMARC
    dmarc_policy = (dmarc_parsed.get("policy", "") if dmarc_parsed else "") or "none"
    narrative = (
        f"Risk Assessment Level: {risk_lvl.capitalize()}\n\n"
        f"Overall result\nDMARC Policy: {dmarc_policy.capitalize()}\n"
        f"Score\n{score}\nof 10\n"
        f"{spf_status}\nSPF\n\nSender Policy Framework\n\n"
        f"{dmarc_status}\nDMARC\n\nDomain-based Message Authentication, Reporting and Conformance\n\n"
        f"{dkim_status}\nDKIM\n\nDomainKeys Identified Mail\n"
    )

    row = {
        "domain": domain,
        "Risk Assessment Level": risk_lvl.capitalize(),
        "DMARC Policy": dmarc_policy,
        "Score": score,
        # Section statuses
        "SPF": spf_status,
        "DMARC": dmarc_status,
        "DKIM": dkim_status,
        # Records / warnings
        "spf_present": "yes" if spf_present else "no",
        "spf_record": spf_rec or "",
        "spf_lookup_count": spf_parsed.get("lookup_count", "") if spf_parsed else "",
        "spf_warnings": " | ".join(spf_w),
        "dmarc_present": "yes" if dmarc_present else "no",
        "dmarc_record": dmarc_rec or "",
        "dmarc_policy": (dmarc_parsed.get("policy", "") if dmarc_parsed else ""),
        "dmarc_pct": (dmarc_parsed.get("pct", "") if dmarc_parsed else ""),
        "dmarc_rua": (dmarc_parsed.get("rua", "") if dmarc_parsed else ""),
        "dmarc_ruf": (dmarc_parsed.get("ruf", "") if dmarc_parsed else ""),
        "dmarc_alignment": f"aspf={dmarc_parsed.get('aspf','')} adkim={dmarc_parsed.get('adkim','')}" if dmarc_parsed else "",
        "dmarc_warnings": " | ".join(dmarc_w),
        "dkim_selectors_found": ",".join(dkim_found),
        "dkim_warnings": " | ".join(dkim_w),
        # Narrative
        "Report": narrative,
        # Helpful link
        "EasyDMARC Link": f"https://easydmarc.com/tools/domain-scanner?domain={domain}",
        "Website": f"https://{domain}",
    }
    return row

# --------------------------- Load domains ---------------------------

def load_domains_hardcoded_excel() -> List[str]:
    """Always load domains from 'domains.xlsx' next to this script (expects a 'Domain' column)."""
    here = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(here, "domains.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"domains.xlsx not found at {xlsx_path}. Create it with a 'Domain' column.", file=sys.stderr)
        return []
    try:
        df = pd.read_excel(xlsx_path)
    except Exception as e:
        print(f"Failed to read {xlsx_path}: {e}", file=sys.stderr)
        return []
    # case-insensitive column match
    col = None
    for c in df.columns:
        if str(c).strip().lower() == "domain":
            col = c
            break
    if not col:
        print("Excel file must contain a 'Domain' column.", file=sys.stderr)
        return []
    domains = [str(d).strip() for d in df[col] if pd.notna(d) and str(d).strip()]
    return domains

# --------------------------- Excel (nice formatting) ---------------------------

def write_excel_nice(merged_df: pd.DataFrame, path: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="Results", index=False)
        ws = writer.book["Results"]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Header style
        header_fill = PatternFill(start_color="FFEEF3FF", end_color="FFEEF3FF", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Column widths (basic autosize heuristic by header length)
        for col in ws.columns:
            max_len = 12
            col_letter = col[0].column_letter
            for cell in col[:200]:  # limit for performance
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_len:
                        max_len = min(len(val), 60)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        # Wrap text for descriptive fields
        wrap_cols = ["SPF Record Description", "Report", "Keywords", "Company Address"]
        for col_name in wrap_cols:
            if col_name in merged_df.columns:
                idx = merged_df.columns.get_loc(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Conditional formatting for Risk and Score
        if "Risk Assessment Level" in merged_df.columns:
            col_idx = merged_df.columns.get_loc("Risk Assessment Level") + 1
            # Simple visual using fill by rule text is tricky; use Score instead for color bands.
        if "Score" in merged_df.columns:
            score_col_idx = merged_df.columns.get_loc("Score") + 1
            # Red for <=3
            ws.conditional_formatting.add(
                f"{ws.cell(row=2, column=score_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=score_col_idx).coordinate}",
                CellIsRule(operator='lessThanOrEqual', formula=['3'], stopIfTrue=True,
                           fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"))
            )
            # Yellow for 4–6
            ws.conditional_formatting.add(
                f"{ws.cell(row=2, column=score_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=score_col_idx).coordinate}",
                CellIsRule(operator='between', formula=['4','6'], stopIfTrue=True,
                           fill=PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"))
            )
            # Green for 7–10
            ws.conditional_formatting.add(
                f"{ws.cell(row=2, column=score_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=score_col_idx).coordinate}",
                CellIsRule(operator='greaterThanOrEqual', formula=['7'], stopIfTrue=True,
                           fill=PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"))
            )

    # Done

# --------------------------- Main ---------------------------

def main():
    ap = argparse.ArgumentParser(description="Scan domains for SPF/DMARC/DKIM and export Excel + CSV.")
    ap.add_argument("--out", "-o", default="results.xlsx", help="Output Excel path (default: results.xlsx)")
    ap.add_argument("--workers", "-w", type=int, default=MAX_WORKERS, help=f"Parallel workers (default: {MAX_WORKERS})")
    args = ap.parse_args()

    domains = load_domains_hardcoded_excel()
    if not domains:
        print("No domains to scan. Ensure 'domains.xlsx' exists with a 'Domain' column.", file=sys.stderr)
        sys.exit(1)

    rows: List[Dict[str, str]] = []
    with futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        for row in ex.map(scan_domain, domains):
            rows.append(row)

    core_df = pd.DataFrame(rows)

    # Build CRM-style DataFrame in the exact column order you provided
    cols = [
        "First Name","Last Name","Title","Company","Email Ready","Called","Company Name for Emails","Email",
        "EasyDMARC Link","SPF","DMARC","DKIM","SPF Record Description","Mobile Phone","Corporate Phone","Other Phone",
        "Industry","Keywords","Person Linkedin Url","Website","Company Linkedin Url","Facebook Url","Twitter Url","City",
        "State","Country","Company Address","Company City","Company State","Company Country","Company Phone","Technologies",
        "Annual Revenue","Total Funding","Latest Funding","Latest Funding Amount","Last Raised At","Subsidiary of","Email Sent",
        "Email Open","Email Bounced","Replied","Demoed","Number of Retail Locations","Apollo Contact Id","Apollo Account Id",
        "Secondary Email","Secondary Email Source","Tertiary Email","Tertiary Email Source","Primary Intent Topic","Primary Intent Score",
        "Secondary Intent Topic","Secondary Intent Score"
    ]

    # Map scanner fields to CRM fields and add additional explicit fields
    crm_df = pd.DataFrame({c: "" for c in cols})
    if not core_df.empty:
        # Copy EasyDMARC-style statuses
        crm_df["EasyDMARC Link"] = core_df["EasyDMARC Link"]
        crm_df["SPF"] = core_df["SPF"]
        crm_df["DMARC"] = core_df["DMARC"]
        crm_df["DKIM"] = core_df["DKIM"]
        crm_df["SPF Record Description"] = core_df["spf_record"].replace({"": "Empty record"})
        # Company / Website
        crm_df["Company Name for Emails"] = core_df["domain"]
        crm_df["Website"] = core_df["Website"]

    # Add explicit fields for clarity (these are extra columns beyond your CRM list)
    explicit_cols = [
        "domain","Risk Assessment Level","DMARC Policy","Score","spf_lookup_count","spf_warnings",
        "dmarc_record","dmarc_rua","dmarc_ruf","dmarc_alignment","dmarc_warnings","dkim_selectors_found","dkim_warnings","Report"
    ]
    explicit_df = core_df[[c for c in explicit_cols if c in core_df.columns]].copy()

    # Merge side-by-side for a single worksheet output
    # Ensure row counts match: repeat/trim to align with domains length
    def align(df_src: pd.DataFrame, n: int) -> pd.DataFrame:
        if df_src.shape[0] == n:
            return df_src.reset_index(drop=True)
        if df_src.empty:
            return pd.DataFrame(index=range(n))
        reps = (n + df_src.shape[0] - 1) // df_src.shape[0]
        return pd.concat([df_src] * reps, ignore_index=True).iloc[:n].reset_index(drop=True)

    n = core_df.shape[0]
    crm_df = align(crm_df, n)
    explicit_df = align(explicit_df, n)

    merged_df = pd.concat([crm_df, explicit_df], axis=1)

    # Write nicely formatted Excel
    write_excel_nice(merged_df, args.out)

    # Also dump a flat CSV of the core scanner results for pipelines
    flat_out_csv = os.path.splitext(args.out)[0] + ".csv"
    core_df.to_csv(flat_out_csv, index=False)

    print(f"Wrote Excel {args.out} and CSV {flat_out_csv} with {n} rows.")

if __name__ == "__main__":
    main()
